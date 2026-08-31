"""تصدير التقارير المحاسبية إلى ملف Excel متعدد الأوراق."""
from pathlib import Path
from datetime import datetime

class ExcelReportExporter:
    def __init__(self, db):
        self.db = db

    def _style(self, ws, headers):
        fill = PatternFill("solid", fgColor="0A2540")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.sheet_view.rightToLeft = True
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(32, len(str(headers[i-1])) + 8))

    def export_all(self, output_dir, from_date=None, to_date=None):
        # تحميل openpyxl عند طلب Excel فقط؛ يسمح ببناء APK الأساسي دون recipe إضافي.
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        output = Path(output_dir); output.mkdir(parents=True, exist_ok=True)
        suffix = f"{from_date or 'all'}_{to_date or 'all'}"
        path = output / f"nexter_reports_{suffix}_{datetime.now():%H%M%S}.xlsx"
        wb = Workbook(); wb.remove(wb.active)
        params=[]; condition=""
        if from_date: condition += " AND substr(issued_at,1,10)>=?"; params.append(from_date)
        if to_date: condition += " AND substr(issued_at,1,10)<=?"; params.append(to_date)
        ws=wb.create_sheet("المبيعات")
        headers=["رقم الفاتورة","التاريخ","النوع","الحالة","قبل الضريبة","الضريبة","الإجمالي","المدفوع"]
        self._style(ws,headers)
        for r in self.db.query(f"SELECT number,issued_at,kind,status,subtotal,tax,total,paid FROM invoices WHERE 1=1{condition} ORDER BY issued_at",params): ws.append([r[h] for h in ("number","issued_at","kind","status","subtotal","tax","total","paid")])
        ws=wb.create_sheet("المصروفات"); headers=["التصنيف","الوصف","المبلغ","الضريبة","التاريخ"]; self._style(ws,headers)
        eparams=[]; ec=""
        if from_date: ec += " AND substr(spent_at,1,10)>=?"; eparams.append(from_date)
        if to_date: ec += " AND substr(spent_at,1,10)<=?"; eparams.append(to_date)
        for r in self.db.query(f"SELECT category,description,amount,tax,spent_at FROM expenses WHERE 1=1{ec} ORDER BY spent_at",eparams): ws.append([r[h] for h in ("category","description","amount","tax","spent_at")])
        ws=wb.create_sheet("المخزون"); headers=["المنتج","الباركود","الوحدة","الكمية","سعر التكلفة","سعر البيع","حد التنبيه"]; self._style(ws,headers)
        for r in self.db.query("SELECT name,barcode,unit,quantity,purchase_price,sale_price,min_quantity FROM products ORDER BY name"): ws.append([r[h] for h in ("name","barcode","unit","quantity","purchase_price","sale_price","min_quantity")])
        ws=wb.create_sheet("ملخص"); self._style(ws,["المؤشر","القيمة"])
        sales=self.db.one("SELECT COALESCE(SUM(total),0) v FROM invoices WHERE kind='sale'")["v"]
        tax=self.db.one("SELECT COALESCE(SUM(tax),0) v FROM invoices WHERE kind='sale'")["v"]
        expenses=self.db.one("SELECT COALESCE(SUM(amount),0) v FROM expenses")["v"]
        for label,value in [("إجمالي المبيعات",sales),("إجمالي الضريبة",tax),("إجمالي المصروفات",expenses),("صافي الربح التقديري",sales-expenses)]: ws.append([label,value])
        wb.save(path); return str(path)
